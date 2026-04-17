"""
research/top_movers/analysis_bundle_builder.py

Builds two companion artifacts per research day:
  1. analysis_bundle_manifest_YYYY-MM-DD.md  — reading guide + artifact table
  2. R1_analysis_bundle_YYYY-MM-DD.xlsx      — all CSV/ledger data in one workbook

xlsx sheets:
  - case_dataset         canonical one-row-per-case (same source as DOCX report)
  - daily_summary        day-level metrics (key/value)
  - signature_candidates today's repeated pattern candidates
  - ledger_snapshot      7-day deduped cross-day view (one row per signature_key)
  - ledger_raw           full raw ledger (all rows, all days)

Downstream-only. Does not touch live runtime, lifecycle, or strategy files.
"""

import csv
import os
from datetime import datetime, timezone
from typing import Dict, List, Optional

from research.top_movers.io import OUTPUT_BASE
from research.top_movers.signature_ledger import (
    LEDGER_PATH,
    load_and_normalize_ledger_rows,
    ledger_rows_as_of,
    build_ledger_snapshot_for_report,
    # Phase 2B
    build_multiday_family_stats,
    build_multiday_interaction_stats,
    build_controlled_validation_summary,
    # Phase 2D
    build_family_validation_stats,
    build_family_answer_contracts,
)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Shared constants
# ---------------------------------------------------------------------------

_HEADER_FILL  = "2E75B6"   # blue
_ALT_FILL     = "F2F2F2"   # light grey for alternating rows
_HEADER_FONT  = "FFFFFF"   # white

# Fields shown in ledger snapshot sheet (same semantics as report snapshot)
_LEDGER_SNAP_FIELDS = [
    "signature_candidate_code", "first_seen_date", "last_seen_date",
    "support_days_count", "recent_support_days_count",
    "latest_validation_status", "current_role",
]

# Fields shown in ledger_raw sheet (key fields from full ledger CSV)
_LEDGER_RAW_FIELDS = [
    "research_day", "signature_candidate_code", "signature_key",
    "support_count_day", "confidence_day", "decision_grade_day",
    "validation_status_day", "first_seen_date", "last_seen_date",
]

# Fields for signature_candidates sheet
_SIG_CAND_FIELDS = [
    "research_day", "signature_candidate_code", "signature_key",
    "support_count", "support_share_pct", "dominant_side",
    "dominant_move_class", "dominant_participation_pattern",
    "dominant_structural_quality", "decision_grade", "confidence",
    "validation_status", "caution_flag", "next_action",
    "maps_to_existing_strategy_family", "improvement_target_layer_mode",
    "median_1h_favor", "median_4h_favor", "notes",
]


# ---------------------------------------------------------------------------
# xlsx helpers
# ---------------------------------------------------------------------------

def _xl_header_fill():
    return PatternFill(fill_type="solid", fgColor=_HEADER_FILL)

def _xl_alt_fill():
    return PatternFill(fill_type="solid", fgColor=_ALT_FILL)

def _xl_header_font():
    return Font(bold=True, color=_HEADER_FONT)

def _xl_write_sheet(ws, headers: List[str], rows: List[List]):
    """Write headers + data rows to a worksheet with basic styling."""
    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _xl_header_fill()
        cell.font = _xl_header_font()
        cell.alignment = Alignment(wrap_text=False)

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        fill = _xl_alt_fill() if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
            if fill:
                cell.fill = fill

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-width (approximate)
    for col_idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        _data_lens = [len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, min(ws.max_row + 1, 52))]
        max_len = max([len(str(header))] + _data_lens) if _data_lens else len(str(header))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def _rows_from_dicts(dicts: List[Dict], fields: List[str]) -> List[List]:
    return [[d.get(f, "") for f in fields] for d in dicts]


def _case_fields(cases: List[Dict]) -> List[str]:
    """Return ordered field list from case rows, preserving canonical insertion order."""
    if not cases:
        return []
    return list(cases[0].keys())


# ---------------------------------------------------------------------------
# Manifest builder
# ---------------------------------------------------------------------------

def build_manifest(
    research_day: str,
    cases: List[Dict],
    sig_candidates: List[Dict],
    daily_summary: Dict,
    output_path: str,
    layer_audit: Optional[dict] = None,
    family_history_snapshot: Optional[dict] = None,
    p2d_family_validation_stats: Optional[List[Dict]] = None,
    p2d_family_answer_contracts: Optional[List[Dict]] = None,
    p2e_controlled_validation_state: Optional[List[Dict]] = None,
    p2e_unseen_day_summary: Optional[List[Dict]] = None,
) -> str:
    """Build analysis_bundle_manifest_YYYY-MM-DD.md"""

    eligible_count  = sum(1 for c in cases if c.get("research_eligible_YN") == "Y")
    health          = daily_summary.get("overall_research_health", "—")
    regime          = daily_summary.get("research_regime", "—")
    improve_count   = daily_summary.get("old_strategy_improvement_cases_count", 0)
    new_thesis      = daily_summary.get("new_strategy_thesis_cases_count", 0)
    short_today     = daily_summary.get("short_intervention_candidate_today", "N")
    short_reason    = daily_summary.get("short_intervention_reason", "none")
    top_sig         = daily_summary.get("top_candidate_signature_1", "none today")

    # Use as-of-day filtered count — same canonical source as unified pack and xlsx bundle.
    # Reading the raw file on disk would include future rows written after this research_day,
    # causing a mismatch with unified pack and xlsx ledger_raw sheet.
    _as_of_ledger = ledger_rows_as_of(research_day, window_days=9999)
    ledger_rows   = len(_as_of_ledger)

    # Phase 2B — derive validation summary for manifest
    _ledger_days_global = len(set(
        r.get("research_day", "") for r in _as_of_ledger if r.get("research_day")
    ))
    _p2b_fam_stats  = build_multiday_family_stats(cases, _as_of_ledger)
    _p2b_inter      = build_multiday_interaction_stats(cases)
    _p2b_val        = build_controlled_validation_summary(
        _p2b_fam_stats, _p2b_inter, _ledger_days_global)
    _p2b_top_family  = _p2b_val.get("top_candidate_family", "—")
    _p2b_promo       = _p2b_val.get("promotion_state", "DESCRIPTIVE_ONLY")
    _p2b_fam_history = _p2b_val.get("family_multiday_history", "not_available_yet")
    _p2b_ledger_days = _p2b_val.get("ledger_research_days_global_context", _ledger_days_global)
    _p2b_bucket      = _p2b_val.get("families_at_bucket_ready_or_above", 0)
    _p2b_tracking    = _p2b_val.get("families_at_tracking_grade_or_above", 0)
    _p2b_recmd       = _p2b_val.get("families_at_recommendation_grade_or_above", 0)
    _p2b_blocking    = _p2b_val.get("blocking_reason", "—")
    _p2b_next        = _p2b_val.get("validation_next_step", "—")
    # Highest case count among SHORT families
    _p2b_top_n       = max(
        (r.get("case_count", 0) for r in _p2b_fam_stats), default=0
    )

    bundle_dir = os.path.dirname(output_path)

    lines = [
        f"# R1 Analysis Bundle Manifest — {research_day}",
        f"",
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        f"",
        f"---",
        f"",
        f"## Day Summary",
        f"",
        f"| Field | Value |",
        f"|---|---|",
        f"| Research Day | {research_day} |",
        f"| Research Regime | {regime} |",
        f"| Overall Health | {health} |",
        f"| Cases Built | {len(cases)} |",
        f"| Eligible Cases | {eligible_count} |",
        f"| Signature Candidates (repeated) | {len(sig_candidates)} |",
        f"| Top Signature | {top_sig} |",
        f"| Improvement Candidates (case-level) | {improve_count} |",
        f"| New Thesis Candidates (case-level) | {new_thesis} |",
        f"| Short Intervention Today | {short_today} |",
        f"| Short Intervention Reason | {short_reason} |",
        f"",
        f"---",
        f"",
        f"## Artifacts in This Bundle",
        f"",
        f"| File | Type | Rows | Description |",
        f"|---|---|---|---|",
        f"| `daily_case_dataset_{research_day}.csv` | canonical case data | {len(cases)} | one-row-per-case; source of truth for all boards |",
        f"| `daily_research_summary_{research_day}.csv` | day summary | 1 | 33+ day-level metrics |",
        f"| `daily_signature_candidates_{research_day}.csv` | repeated patterns | {len(sig_candidates)} | patterns with ≥2 eligible cases |",
        f"| `signature_evidence_ledger.csv` | rolling ledger | {ledger_rows} | cross-day evidence (as of {research_day}) |",
        f"| `R1_analysis_bundle_{research_day}.xlsx` | xlsx bundle | — | all above in one workbook |",
        f"",
        f"---",
        f"",
        f"## Phase 2B Multi-Day Measurement Summary",
        f"",
        f"_Phase 2B = today-only conservative promotion frame. "
        f"Top Family and multiday history are shown in Phase 2C section below._",
        f"",
        f"| Field | Value |",
        f"|---|---|",
        f"| Promotion State | {_p2b_promo} |",
        f"| Families at Bucket-Ready or Above (>= 10) | {_p2b_bucket} |",
        f"| Families at Tracking-Grade or Above (>= 20) | {_p2b_tracking} |",
        f"| Families at Recommendation-Grade or Above (>= 50) | {_p2b_recmd} |",
        f"| Blocking Reason | {_p2b_blocking} |",
        f"| Validation Next Step | {_p2b_next} |",
        f"",
        f"> **Note:** RECOMMENDATION_GRADE requires >= 50 cases. "
        f"TRACKING_GRADE requires >= 20. "
        f"Do not read promotion_state as action-ready below PREPARE_HYPOTHESIS.",
        f"",
        f"---",
        f"",
        f"## How to Use This Bundle",
        f"",
        f"Send to ChatGPT in this order for best context:",
        f"1. This manifest file",
        f"2. `R1_analysis_bundle_{research_day}.xlsx` (or the individual CSVs)",
        f"",
        f"---",
        f"",
        f"## Critical Semantic Rules (read before interpreting)",
        f"",
        f"| Rule | Clarification |",
        f"|---|---|",
        f"| repeated signatures ≠ strategy proof | A repeated signature = ≥2 eligible cases share the same (side, pre_move_sig, participation, structural_quality). It is a pattern to track, NOT a validated strategy or a live rule change recommendation. |",
        f"| case-level theses ≠ repeated patterns | A case can have decision_grade = NEW_STRATEGY_THESIS_CANDIDATE even when zero repeated signatures exist today. These are per-case conclusions, independent of the signature evidence board. |",
        f"| research families under investigation ≠ live strategy families | R1 research families (long_breakout_retest, short_exhaustion_retest) are the mapping targets. Do not confuse NEW_STRATEGY_THESIS_CANDIDATE cases with changes to the live bot strategy families. |",
        f"| CLEAN_WITH_VISUAL_GAPS health | Research data is analysis-ready (eligible/proxy/outcome complete) but ≥1 chart image could not be rendered. Outcome fields are unaffected — they come from price data. Do not treat as PARTIAL. |",
        f"| data_confidence = case data trust | HIGH/MEDIUM/LOW trust in the completeness and interpretability of this case's data inputs. |",
        f"| intervention_confidence = intervention readiness | Whether this case is credible evidence for an intervention discussion. Separate from data quality. |",
        f"| anchor_conflict_flag = Y | One or more anchors used a fallback detection (no clean signal found). Treat anchor-derived features with caution for these cases. |",
        f"",
        f"## Readiness Levels",
        f"",
        f"| Level | Meaning |",
        f"|---|---|",
        f"| descriptive_only | Not enough repetition or strategy relevance. Observe only. |",
        f"| keep_tracking | Interesting but insufficient for any intervention. |",
        f"| old_strategy_improvement_candidate | Points to an existing strategy family + identifiable improvement layer. Requires further validation. |",
        f"| new_strategy_thesis_candidate | Potential new family, not yet validated. Needs multi-day repeated support before any action. |",
        f"",
        f"---",
        f"",
        f"## Sheet Guide (xlsx)",
        f"",
        f"| Sheet | Contents | Use |",
        f"|---|---|---|",
        f"| case_dataset | All {len(cases)} cases with full field set | Primary source of truth — trace any board row back here |",
        f"| daily_summary | 33+ day-level fields | Day health, dominant patterns, short intervention flag |",
        f"| signature_candidates | {len(sig_candidates)} repeated pattern candidates | Cross-case pattern evidence |",
        f"| ledger_snapshot | 7-day deduped view (one row per signature_key) | Cross-day tracking |",
        f"| ledger_raw | Raw ledger rows as of {research_day} | Audit trail up to report day |",
        f"| multiday_family_stats | SHORT family stats — today-only case counts (family_days_count='not_available_yet' by design) | Phase 2B SHORT-side evidence |",
        f"| multiday_interaction_stats | 4 interaction pairs (today's SHORT cases) | Phase 2B cross-dimension |",
        f"| validation_summary | Promotion state + blocking reason + global ledger context. phase_2b_family_multiday_history key = not_available_yet — multiday truth is in family_case_history sheet | Phase 2B controlled validation |",
        f"| layer_field_coverage_audit | Layer 0–8 mandatory field coverage vs master spec | Phase 2C field gap analysis |",
        f"| family_case_history | SHORT family cross-day case history (from case datasets — NOT from ledger) | Phase 2C family history foundation |",
        f"| family_validation_stats | Phase 2D statistical validation per family: bootstrap CI, KS/t-test, sample grades | Phase 2D statistical validation |",
        f"| family_answer_contracts | Phase 2D answer contracts for 6 question families (Exhaustion / Breakdown / Retest / Timing / Invalidation / Context) | Phase 2D family answer contracts |",
        f"| controlled_validation_state | Phase 2E gate state per display_family: DESCRIPTIVE_ONLY / KEEP_TRACKING / NOT_READY / PREPARE_HYPOTHESIS / READY_FOR_CONTROLLED_VALIDATION. Includes anchor_qa_gate and gating_reasons. | Phase 2E promotion gate |",
        f"| unseen_day_validation | Phase 2E holdout/unseen-day validation summary per family. First pass = NOT_READY_no_holdout_data for all. | Phase 2E unseen-day gate |",
        f"",
        f"---",
        f"",
        f"*This manifest is downstream-only. It does not modify live bot config, strategy, lifecycle, or runtime files.*",
    ]

    # --- Phase 2C Foundation Summary ---
    if layer_audit or family_history_snapshot:
        la   = layer_audit or {}
        fhs  = family_history_snapshot or {}
        _blk = la.get("blocking_layers", [])
        _missing = la.get("fields_missing_count", "—")
        _null    = la.get("fields_all_null_count", "—")
        _total   = la.get("total_required_fields", "—")
        _fh_status   = fhs.get("family_history_status", "not_available_yet")
        _fh_top      = fhs.get("top_candidate_family", "—")
        _fh_multi    = fhs.get("top_family_case_count_multiday", "—")
        _fh_days     = fhs.get("top_family_days_count", "—")
        _fh_fams     = fhs.get("families_with_history_count", "—")
        _fh_hist_days= fhs.get("historical_case_days_loaded", "—")
        _fh_window   = fhs.get("history_window_days", "—")
        # _blocker: reflect actual current state, not pre-Phase-2D wording
        _p2d_present = (
            p2d_family_validation_stats is not None
            or p2d_family_answer_contracts is not None
        )
        _p2e_present = (
            p2e_controlled_validation_state is not None
            or p2e_unseen_day_summary is not None
        )
        if _fh_status == "available":
            if _p2e_present:
                _blocker = (
                    "Phase 2D stats exist. PREPARE_HYPOTHESIS blocked by Phase 2E gates "
                    "(anchor QA, holdout data, sample thresholds) — see Phase 2E section."
                )
            elif _p2d_present:
                _blocker = (
                    "Phase 2D stats exist. Phase 2E gate required before any promotion."
                )
            else:
                _blocker = (
                    "Family history exists. Phase 2D statistical validation required next."
                )
        else:
            _blocker = "Family history foundation not yet built — run historical days."
        lines += [
            f"",
            f"---",
            f"",
            f"## Phase 2C Foundation Summary",
            f"",
            f"### Layer 0–8 Field Coverage",
            f"",
            f"| Field | Value |",
            f"|---|---|",
            f"| Layer Coverage Audit Status | {la.get('blocking_layers_count', '—')} blocking layer(s) of {la.get('layers_checked','—')} |",
            f"| Total Mandatory Fields Checked | {_total} |",
            f"| Missing Mandatory Fields | {_missing} |",
            f"| All-Null Mandatory Fields | {_null} |",
            f"| Blocking Layers | {_blk if _blk else 'none'} |",
            f"",
            f"### Family History Foundation",
            f"",
            f"| Field | Value |",
            f"|---|---|",
            f"| Historical Case Days Loaded | {_fh_hist_days} |",
            f"| History Window Days | {_fh_window} |",
            f"| Families With Cross-Day History | {_fh_fams} |",
            f"| Top SHORT Family (multiday) | {_fh_top} |",
            f"| Top SHORT Family Case Count (multiday) | {_fh_multi} |",
            f"| Top SHORT Family Days | {_fh_days} |",
            f"| Family History Status | {_fh_status} |",
            f"| Current Blocker Before Stats | {_blocker} |",
            f"",
            f"> Family history sourced from daily_case_dataset CSVs — NOT from signature_evidence_ledger. "
            f"PREPARE_HYPOTHESIS remains blocked in Phase 2C.",
        ]

    # --- Phase 2D Statistical Validation Summary ---
    if p2d_family_validation_stats is not None or p2d_family_answer_contracts is not None:
        _p2d_val  = p2d_family_validation_stats or []
        _p2d_ctr  = p2d_family_answer_contracts or []
        _best_n         = max((r.get("case_count_multiday", 0) for r in _p2d_val), default=0)
        # Largest bucket (any family including unclassified)
        _largest_bucket = next(
            (r.get("display_family", "—") for r in _p2d_val
             if r.get("case_count_multiday", 0) == _best_n),
            "—"
        )
        # Best named/actionable family (excludes unclassified)
        _best_named_n   = max(
            (r.get("case_count_multiday", 0) for r in _p2d_val
             if r.get("is_actionable_family", True)),
            default=0,
        )
        _best_named_fam = next(
            (r.get("display_family", "—") for r in _p2d_val
             if r.get("is_actionable_family", True)
             and r.get("case_count_multiday", 0) == _best_named_n),
            "—"
        )
        _blocked_ctr = sum(1 for r in _p2d_ctr if r.get("recommendation_state") == "blocked")
        _kt_ctr      = sum(1 for r in _p2d_ctr if r.get("recommendation_state") == "keep_tracking")
        lines += [
            f"",
            f"---",
            f"",
            f"## Phase 2D Statistical Validation Summary",
            f"",
            f"_Phase 2D = statistical validation layer. Conservative by design. "
            f"PREPARE_HYPOTHESIS not unlocked here._",
            f"",
            f"### Family Validation Stats",
            f"",
            f"| Field | Value |",
            f"|---|---|",
            f"| Families Analyzed | {len(_p2d_val)} |",
            f"| Largest Bucket (any family) | {_largest_bucket} ({_best_n} cases) |",
            f"| Best Named / Actionable Family | {_best_named_fam} ({_best_named_n} cases) |",
            f"| Note | See family_validation_stats sheet for bootstrap CI, KS/t-test, sample grades |",
            f"",
            f"### Answer Contracts (6 Question Families)",
            f"",
            f"| Family | Recommendation State | Blocking Reason |",
            f"|---|---|---|",
        ]
        for r in _p2d_ctr:
            def _tw(text, limit=90):
                s = str(text)
                if len(s) <= limit:
                    return s
                return s[:limit].rsplit(" ", 1)[0] + "…"
            lines.append(
                f"| {r.get('family_name','—')} "
                f"| {r.get('recommendation_state','—')} "
                f"| {_tw(r.get('blocking_reason','—'))} |"
            )
        lines += [
            f"",
            f"Contracts blocked: {_blocked_ctr}/6. "
            f"Keep-tracking: {_kt_ctr}/6. "
            f"See family_answer_contracts sheet for full contract detail.",
            f"",
            f"> Phase 2D does not write live-rule instructions. "
            f"Recommendation state remains conservative.",
        ]

    # --- Phase 2E Controlled Validation Gate ---
    if p2e_controlled_validation_state is not None or p2e_unseen_day_summary is not None:
        _cvs = p2e_controlled_validation_state or []
        _uds = p2e_unseen_day_summary or []
        _promoted = [
            r for r in _cvs
            if r.get("controlled_validation_state") in
               ("PREPARE_HYPOTHESIS", "READY_FOR_CONTROLLED_VALIDATION")
        ]
        _not_ready = [
            r for r in _cvs
            if r.get("controlled_validation_state") ==
               "NOT_READY_FOR_CONTROLLED_VALIDATION"
        ]

        def _tw90(text, limit=130):
            s = str(text)
            return s if len(s) <= limit else s[:limit - 1].rsplit(" ", 1)[0] + "\u2026"

        lines += [
            f"",
            f"---",
            f"",
            f"## Phase 2E \u2014 Controlled Validation Gate",
            f"",
            f"_Gate-building phase. Conservative result expected at current evidence level._",
            f"",
            f"| Field | Value |",
            f"|---|---|",
            f"| Families Evaluated | {len(_cvs)} |",
            f"| Families Promoted (PREPARE_HYPOTHESIS+) | {len(_promoted)} |",
            f"| Families NOT_READY | {len(_not_ready)} |",
            f"| Unseen-Day Summary Rows | {len(_uds)} |",
            f"| Unseen-Day Status (first row) | "
            f"{_uds[0].get('unseen_validation_status', '\u2014') if _uds else '\u2014'} |",
            f"",
            f"| Family | State | N Multi | Anchor Gate | Blocker Summary |",
            f"|---|---|---|---|---|",
        ]
        for r in _cvs:
            lines.append(
                f"| {r.get('display_family', '\u2014')} "
                f"| {r.get('controlled_validation_state', '\u2014')} "
                f"| {r.get('case_count_multiday', '\u2014')} "
                f"| {r.get('anchor_qa_gate', '\u2014')} "
                f"| {_tw90(r.get('promotion_blocker_summary', '\u2014'))} |"
            )
        lines += [
            f"",
            f"> Promotion is blocked for all families at current evidence level. "
            f"See controlled_validation_state sheet in xlsx bundle for full detail.",
        ]

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    return output_path


# ---------------------------------------------------------------------------
# xlsx bundle builder
# ---------------------------------------------------------------------------

def build_xlsx_bundle(
    research_day: str,
    cases: List[Dict],
    sig_candidates: List[Dict],
    daily_summary: Dict,
    output_path: str,
    window_days: int = 7,
    layer_audit: Optional[dict] = None,
    family_case_history: Optional[List[Dict]] = None,
    p2d_family_validation_stats: Optional[List[Dict]] = None,
    p2d_family_answer_contracts: Optional[List[Dict]] = None,
    p2e_controlled_validation_state: Optional[List[Dict]] = None,
    p2e_unseen_day_summary: Optional[List[Dict]] = None,
) -> str:
    """
    Build R1_analysis_bundle_YYYY-MM-DD.xlsx with up to 14 sheets.
    Requires openpyxl: pip install openpyxl
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for xlsx bundle: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # ------------------------------------------------------------------
    # Sheet 1: case_dataset
    # ------------------------------------------------------------------
    ws_cases = wb.create_sheet("case_dataset")
    if cases:
        fields = _case_fields(cases)
        rows   = _rows_from_dicts(cases, fields)
        _xl_write_sheet(ws_cases, fields, rows)
    else:
        ws_cases.cell(row=1, column=1, value="No cases for this day")

    # ------------------------------------------------------------------
    # Sheet 2: daily_summary (key/value layout)
    # ------------------------------------------------------------------
    ws_summary = wb.create_sheet("daily_summary")
    _xl_write_sheet(ws_summary, ["Field", "Value"],
                    [[k, str(v) if v is not None else ""] for k, v in daily_summary.items()])

    # ------------------------------------------------------------------
    # Sheet 3: signature_candidates
    # ------------------------------------------------------------------
    ws_sigs = wb.create_sheet("signature_candidates")
    if sig_candidates:
        # Use predefined display fields; fall back to all fields if some missing
        available = set(sig_candidates[0].keys())
        display_fields = [f for f in _SIG_CAND_FIELDS if f in available]
        extra = [f for f in sig_candidates[0].keys() if f not in display_fields]
        all_fields = display_fields + extra
        rows = _rows_from_dicts(sig_candidates, all_fields)
        _xl_write_sheet(ws_sigs, all_fields, rows)
    else:
        ws_sigs.cell(row=1, column=1, value="No repeated signature candidates today")
        ws_sigs.cell(row=2, column=1, value="(zero candidates = no cross-case pattern met threshold — valid)")

    # ------------------------------------------------------------------
    # Sheet 4: ledger_snapshot (7-day deduped, one row per signature_key)
    # ------------------------------------------------------------------
    ws_snap = wb.create_sheet("ledger_snapshot")
    # Canonical as-of-day source — same derivation rules as unified pack
    normalized = ledger_rows_as_of(research_day, window_days=window_days)
    snapshot   = build_ledger_snapshot_for_report(normalized, research_day)
    if snapshot:
        snap_fields = _LEDGER_SNAP_FIELDS + [
            f for f in snapshot[0].keys()
            if f not in _LEDGER_SNAP_FIELDS and not f.startswith("_")
        ]
        rows = _rows_from_dicts(snapshot, snap_fields)
        _xl_write_sheet(ws_snap, snap_fields, rows)
    else:
        ws_snap.cell(row=1, column=1, value="No ledger data in rolling window")
        ws_snap.cell(row=2, column=1, value=f"(window: {window_days} days ending {research_day})")

    # ------------------------------------------------------------------
    # Sheet 5: ledger_raw (full raw ledger — all days)
    # ------------------------------------------------------------------
    ws_raw = wb.create_sheet("ledger_raw")
    # Use same canonical source as ledger_snapshot — no separate disk read
    raw_rows = list(normalized)  # already filtered by ledger_rows_as_of above
    if raw_rows:
        # Show key fields first, then remaining
        available = set(raw_rows[0].keys())
        key_first = [f for f in _LEDGER_RAW_FIELDS if f in available]
        rest = [f for f in raw_rows[0].keys() if f not in key_first]
        all_fields = key_first + rest
        sorted_rows = sorted(raw_rows, key=lambda r: r.get("research_day", ""), reverse=True)
        rows = _rows_from_dicts(sorted_rows, all_fields)
        _xl_write_sheet(ws_raw, all_fields, rows)
    else:
        ws_raw.cell(row=1, column=1, value="No ledger data as of this report day")

    # ------------------------------------------------------------------
    # Sheet 6: multiday_family_stats (Phase 2B)
    # ------------------------------------------------------------------
    ws_fam = wb.create_sheet("multiday_family_stats")
    # Consume helpers from signature_ledger — no re-derive here
    p2b_fam_stats = build_multiday_family_stats(cases, normalized)
    if p2b_fam_stats:
        fam_fields = list(p2b_fam_stats[0].keys())
        fam_rows   = _rows_from_dicts(p2b_fam_stats, fam_fields)
        _xl_write_sheet(ws_fam, fam_fields, fam_rows)
    else:
        ws_fam.cell(row=1, column=1, value="No eligible cases for family stats today")
        ws_fam.cell(row=2, column=1,
                    value="(all cases excluded or no research_eligible_YN=Y rows)")

    # ------------------------------------------------------------------
    # Sheet 7: multiday_interaction_stats (Phase 2B)
    # ------------------------------------------------------------------
    ws_inter = wb.create_sheet("multiday_interaction_stats")
    p2b_inter_stats = build_multiday_interaction_stats(cases)
    if p2b_inter_stats:
        # Flatten: one row per combination (not per pair)
        flat_inter_rows = []
        for pair in p2b_inter_stats:
            for combo in pair.get("top_combinations", []):
                flat_inter_rows.append(combo)
        if flat_inter_rows:
            inter_fields = list(flat_inter_rows[0].keys())
            inter_rows   = _rows_from_dicts(flat_inter_rows, inter_fields)
            _xl_write_sheet(ws_inter, inter_fields, inter_rows)
        else:
            ws_inter.cell(row=1, column=1, value="No interaction combinations found today")
    else:
        ws_inter.cell(row=1, column=1, value="No eligible cases for interaction stats today")

    # ------------------------------------------------------------------
    # Sheet 8: validation_summary (Phase 2B)
    # ------------------------------------------------------------------
    ws_val = wb.create_sheet("validation_summary")
    _p2b_ledger_days_ctx = len(set(
        r.get("research_day", "") for r in normalized if r.get("research_day")
    ))
    p2b_val_summary = build_controlled_validation_summary(
        p2b_fam_stats, p2b_inter_stats, _p2b_ledger_days_ctx)
    # Render as key/value — exclude internal private fields (prefixed with _).
    # Gate 0 — rename family_multiday_history key for xlsx semantic clarity.
    # The value stays "not_available_yet": Phase 2B is today-only by design.
    # Phase 2C multiday truth lives in the separate family_case_history sheet.
    _VAL_RENAME_P2B = {
        "family_multiday_history": (
            "phase_2b_family_multiday_history"
            " (see family_case_history sheet for Phase 2C multiday data)"
        ),
    }
    _xl_write_sheet(ws_val, ["Field", "Value"],
                    [[_VAL_RENAME_P2B.get(k, k), str(v) if v is not None else ""]
                     for k, v in p2b_val_summary.items()
                     if not k.startswith("_")])

    # ------------------------------------------------------------------
    # Sheet 9: layer_field_coverage_audit (Phase 2C)
    # ------------------------------------------------------------------
    ws_audit = wb.create_sheet("layer_field_coverage_audit")
    if layer_audit and layer_audit.get("layer_rows"):
        audit_fields = list(layer_audit["layer_rows"][0].keys())
        audit_rows   = _rows_from_dicts(layer_audit["layer_rows"], audit_fields)
        _xl_write_sheet(ws_audit, audit_fields, audit_rows)
    else:
        ws_audit.cell(row=1, column=1, value="Layer audit not available or not provided")
        ws_audit.cell(row=2, column=1, value="Run with layer_audit param to populate")

    # ------------------------------------------------------------------
    # Sheet 10: family_case_history (Phase 2C)
    # ------------------------------------------------------------------
    ws_hist = wb.create_sheet("family_case_history")
    if family_case_history:
        hist_fields = list(family_case_history[0].keys())
        hist_rows   = _rows_from_dicts(family_case_history, hist_fields)
        _xl_write_sheet(ws_hist, hist_fields, hist_rows)
    else:
        ws_hist.cell(row=1, column=1, value="Family case history not available")
        ws_hist.cell(row=2, column=1, value="(no historical case datasets found in window, or not provided)")

    # ------------------------------------------------------------------
    # Sheet 11: family_validation_stats (Phase 2D)
    # ------------------------------------------------------------------
    ws_p2d_val = wb.create_sheet("family_validation_stats")
    if p2d_family_validation_stats:
        p2d_val_fields = list(p2d_family_validation_stats[0].keys())
        p2d_val_rows   = _rows_from_dicts(p2d_family_validation_stats, p2d_val_fields)
        _xl_write_sheet(ws_p2d_val, p2d_val_fields, p2d_val_rows)
    else:
        ws_p2d_val.cell(row=1, column=1, value="Phase 2D family validation stats not available")
        ws_p2d_val.cell(row=2, column=1,
                        value="(build_family_validation_stats returned empty — "
                              "run after historical case datasets exist)")

    # ------------------------------------------------------------------
    # Sheet 12: family_answer_contracts (Phase 2D)
    # ------------------------------------------------------------------
    ws_p2d_ctr = wb.create_sheet("family_answer_contracts")
    if p2d_family_answer_contracts:
        p2d_ctr_fields = list(p2d_family_answer_contracts[0].keys())
        p2d_ctr_rows   = _rows_from_dicts(p2d_family_answer_contracts, p2d_ctr_fields)
        _xl_write_sheet(ws_p2d_ctr, p2d_ctr_fields, p2d_ctr_rows)
    else:
        ws_p2d_ctr.cell(row=1, column=1, value="Phase 2D family answer contracts not available")
        ws_p2d_ctr.cell(row=2, column=1,
                        value="(build_family_answer_contracts returned empty — "
                              "run after historical case datasets exist)")

    # ------------------------------------------------------------------
    # Sheet 13: controlled_validation_state (Phase 2E)
    # ------------------------------------------------------------------
    ws_p2e_cvs = wb.create_sheet("controlled_validation_state")
    if p2e_controlled_validation_state:
        _cvs_flat = []
        for row in p2e_controlled_validation_state:
            flat_row = dict(row)
            reasons = flat_row.get("gating_reasons", [])
            flat_row["gating_reasons"] = " | ".join(str(r) for r in reasons)
            _cvs_flat.append(flat_row)
        cvs_fields = list(_cvs_flat[0].keys())
        cvs_rows   = _rows_from_dicts(_cvs_flat, cvs_fields)
        _xl_write_sheet(ws_p2e_cvs, cvs_fields, cvs_rows)
    else:
        ws_p2e_cvs.cell(row=1, column=1,
                        value="Phase 2E controlled validation state not available")
        ws_p2e_cvs.cell(row=2, column=1,
                        value="(build_controlled_validation_state returned empty "
                              "\u2014 run after Phase 2C family history + Phase 2D stats exist)")

    # ------------------------------------------------------------------
    # Sheet 14: unseen_day_validation (Phase 2E)
    # ------------------------------------------------------------------
    ws_p2e_uds = wb.create_sheet("unseen_day_validation")
    if p2e_unseen_day_summary:
        uds_fields = list(p2e_unseen_day_summary[0].keys())
        uds_rows   = _rows_from_dicts(p2e_unseen_day_summary, uds_fields)
        _xl_write_sheet(ws_p2e_uds, uds_fields, uds_rows)
    else:
        ws_p2e_uds.cell(row=1, column=1,
                        value="Phase 2E unseen-day validation summary not available")
        ws_p2e_uds.cell(row=2, column=1,
                        value="(build_unseen_day_validation_summary returned empty "
                              "\u2014 run after Phase 2C family history exists)")

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path
